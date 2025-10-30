from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for
import difflib
import re
import os
import sqlite3
import hashlib
import json
from datetime import datetime
from collections import Counter
import math
from werkzeug.utils import secure_filename
import PyPDF2
from docx import Document
import openpyxl
from PIL import Image
import pytesseract

# Avoid importing language_tool_python at module import time to prevent startup hangs
LANGUAGE_TOOL_AVAILABLE = False  # Will be updated lazily inside check_grammar if enabled
ENABLE_GRAMMAR_TOOL = os.getenv('ENABLE_GRAMMAR_TOOL', '0') in ('1', 'true', 'True', 'yes', 'YES')

try:
    import textstat
    TEXTSTAT_AVAILABLE = True
except ImportError:
    TEXTSTAT_AVAILABLE = False
    print("Warning: textstat not available. Readability analysis will be limited.")

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

ALLOWED_EXTENSIONS = {
    'txt', 'pdf', 'docx', 'doc', 'xlsx', 'xls', 'png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff'
}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def init_db():
    conn = sqlite3.connect('plagiarism_detector.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            file_hash TEXT UNIQUE NOT NULL,
            file_type TEXT NOT NULL,
            content TEXT NOT NULL,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            file_size INTEGER
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_name TEXT NOT NULL,
            doc1_name TEXT NOT NULL,
            doc2_name TEXT NOT NULL,
            similarity_percentage REAL NOT NULL,
            report_data TEXT NOT NULL,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS single_file_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_name TEXT NOT NULL,
            document_id INTEGER NOT NULL,
            report_data TEXT NOT NULL,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (document_id) REFERENCES documents (id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS batch_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_name TEXT NOT NULL,
            report_data TEXT NOT NULL,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_hash(file_path):
    with open(file_path, 'rb') as f:
        return hashlib.md5(f.read()).hexdigest()

def extract_text_from_file(file_path, file_type):
    try:
        if file_type == 'txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif file_type == 'pdf':
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text()
                return text
        elif file_type in ['docx', 'doc']:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        elif file_type in ['xlsx', 'xls']:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text += str(cell.value) + " "
                    text += "\n"
            return text
        elif file_type in ['png', 'jpg', 'jpeg', 'gif', 'bmp', 'tiff']:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text
        else:
            return "Error: Unsupported file type"
    except Exception as e:
        return f"Error extracting text: {str(e)}"

def save_document_to_db(filename, file_path, file_type, content):
    conn = sqlite3.connect('plagiarism_detector.db')
    cursor = conn.cursor()
    file_hash = get_file_hash(file_path)
    file_size = os.path.getsize(file_path)
    try:
        cursor.execute('''
            INSERT INTO documents (filename, file_hash, file_type, content, file_size)
            VALUES (?, ?, ?, ?, ?)
        ''', (filename, file_hash, file_type, content, file_size))
        doc_id = cursor.lastrowid
        conn.commit()
        return doc_id
    except sqlite3.IntegrityError:
        cursor.execute('SELECT id FROM documents WHERE file_hash = ?', (file_hash,))
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else None
    finally:
        conn.close()

def preprocess_text(text):
    text = text.lower()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def calculate_cosine_similarity(text1, text2):
    text1 = preprocess_text(text1)
    text2 = preprocess_text(text2)
    words1 = set(text1.split())
    words2 = set(text2.split())
    all_words = words1.union(words2)
    vector1 = [1 if word in words1 else 0 for word in all_words]
    vector2 = [1 if word in words2 else 0 for word in all_words]
    dot_product = sum(a * b for a, b in zip(vector1, vector2))
    magnitude1 = math.sqrt(sum(a * a for a in vector1))
    magnitude2 = math.sqrt(sum(b * b for b in vector2))
    if magnitude1 == 0 or magnitude2 == 0:
        return 0
    return dot_product / (magnitude1 * magnitude2)

def calculate_sequence_similarity(text1, text2):
    return difflib.SequenceMatcher(None, text1.lower(), text2.lower()).ratio()

def detect_plagiarism(text1, text2, doc1_name="Document 1", doc2_name="Document 2"):
    cosine_sim = calculate_cosine_similarity(text1, text2)
    sequence_sim = calculate_sequence_similarity(text1, text2)
    similarity = (cosine_sim * 0.6) + (sequence_sim * 0.4)
    percentage = round(similarity * 100, 2)
    if percentage >= 80:
        status, color = "High Plagiarism Risk", "danger"
    elif percentage >= 60:
        status, color = "Moderate Plagiarism Risk", "warning"
    elif percentage >= 30:
        status, color = "Low Plagiarism Risk", "info"
    else:
        status, color = "Original Content", "success"
    analysis = {
        'word_count_1': len(text1.split()),
        'word_count_2': len(text2.split()),
        'char_count_1': len(text1),
        'char_count_2': len(text2),
        'common_words': find_common_words(text1, text2),
        'unique_words_1': find_unique_words(text1, text2),
        'unique_words_2': find_unique_words(text2, text1),
        'similar_sentences': find_similar_sentences(text1, text2),
        'risk_level': get_risk_level(percentage),
        'recommendations': get_recommendations(percentage)
    }
    return {
        'similarity': similarity,
        'percentage': percentage,
        'cosine_similarity': round(cosine_sim * 100, 2),
        'sequence_similarity': round(sequence_sim * 100, 2),
        'status': status,
        'color': color,
        'details': f"Similarity between {doc1_name} and {doc2_name}",
        'analysis': analysis
    }

def find_common_words(text1, text2):
    words1 = set(preprocess_text(text1).split())
    words2 = set(preprocess_text(text2).split())
    return list(words1.intersection(words2))[:20]

def find_unique_words(text1, text2):
    words1 = set(preprocess_text(text1).split())
    words2 = set(preprocess_text(text2).split())
    return list(words1 - words2)[:20]

def find_similar_sentences(text1, text2):
    sentences1 = [s.strip() for s in text1.split('.') if s.strip()]
    sentences2 = [s.strip() for s in text2.split('.') if s.strip()]
    similar_sentences = []
    for s1 in sentences1:
        for s2 in sentences2:
            if len(s1) > 20 and len(s2) > 20:
                similarity = difflib.SequenceMatcher(None, s1.lower(), s2.lower()).ratio()
                if similarity > 0.5:
                    similar_sentences.append({
                        'sentence1': s1,
                        'sentence2': s2,
                        'similarity': round(similarity * 100, 2)
                    })
    return similar_sentences[:10]

def get_risk_level(percentage):
    if percentage >= 80:
        return "CRITICAL"
    elif percentage >= 60:
        return "HIGH"
    elif percentage >= 30:
        return "MEDIUM"
    else:
        return "LOW"

def get_recommendations(percentage):
    if percentage >= 80:
        return [
            "This content shows very high similarity. Consider complete rewriting.",
            "Review and cite sources properly if this is research work.",
            "Ensure proper attribution for any quoted material."
        ]
    elif percentage >= 60:
        return [
            "Moderate similarity detected. Review similar sections.",
            "Consider paraphrasing similar content.",
            "Add proper citations where needed."
        ]
    elif percentage >= 30:
        return [
            "Low similarity detected. This is generally acceptable.",
            "Review highlighted similar sections if needed.",
            "Ensure proper citation practices."
        ]
    else:
        return [
            "Content appears to be original.",
            "Continue with good writing practices.",
            "Always cite sources when using external material."
        ]

def check_grammar(text):
    grammar_issues = []
    readability_scores = {}

    if ENABLE_GRAMMAR_TOOL:
        try:
            global LANGUAGE_TOOL_AVAILABLE
            if not LANGUAGE_TOOL_AVAILABLE:
                import importlib
                language_tool_python = importlib.import_module('language_tool_python')
                LANGUAGE_TOOL_AVAILABLE = True
            tool = language_tool_python.LanguageTool('en-US')
            matches = tool.check(text)
            for match in matches:
                issue = {
                    'message': match.message,
                    'context': match.context,
                    'offset': match.offset,
                    'length': match.length,
                    'rule_id': match.ruleId,
                    'replacements': match.replacements[:3] if match.replacements else []
                }
                grammar_issues.append(issue)
            tool.close()
        except Exception as e:
            print(f"LanguageTool error: {e}")
            grammar_issues = basic_grammar_check(text)
    else:
        grammar_issues = basic_grammar_check(text)

    if TEXTSTAT_AVAILABLE:
        try:
            readability_scores = {
                'flesch_reading_ease': textstat.flesch_reading_ease(text),
                'flesch_kincaid_grade': textstat.flesch_kincaid_grade(text),
                'gunning_fog': textstat.gunning_fog(text),
                'smog_index': textstat.smog_index(text),
                'automated_readability_index': textstat.automated_readability_index(text)
            }
        except Exception as e:
            print(f"Textstat error: {e}")
            readability_scores = basic_readability_analysis(text)
    else:
        readability_scores = basic_readability_analysis(text)
    text_stats = {
        'word_count': len(text.split()),
        'sentence_count': len([s for s in text.split('.') if s.strip()]),
        'syllable_count': estimate_syllables(text),
        'character_count': len(text)
    }
    return {
        'grammar_issues': grammar_issues,
        'total_issues': len(grammar_issues),
        'readability_scores': readability_scores,
        'text_stats': text_stats
    }

def basic_grammar_check(text):
    issues = []
    sentences = text.split('.')
    for i, sentence in enumerate(sentences):
        sentence = sentence.strip()
        if not sentence:
            continue
        if '  ' in sentence:
            issues.append({
                'message': 'Double space detected',
                'context': sentence,
                'offset': 0,
                'length': len(sentence),
                'rule_id': 'DOUBLE_SPACE',
                'replacements': [sentence.replace('  ', ' ')]
            })
        if i > 0 and sentence and sentence[0].islower():
            issues.append({
                'message': 'Sentence should start with capital letter',
                'context': sentence,
                'offset': 0,
                'length': 1,
                'rule_id': 'CAPITALIZATION',
                'replacements': [sentence[0].upper() + sentence[1:]]
            })
        common_mistakes = {
            'teh': 'the',
            'adn': 'and',
            'recieve': 'receive',
            'seperate': 'separate',
            'occured': 'occurred',
            'definately': 'definitely',
            'accomodate': 'accommodate',
            'begining': 'beginning',
            'neccessary': 'necessary',
            'occassion': 'occasion'
        }
        for mistake, correction in common_mistakes.items():
            if mistake in sentence.lower():
                issues.append({
                    'message': f'Possible misspelling: "{mistake}"',
                    'context': sentence,
                    'offset': 0,
                    'length': len(sentence),
                    'rule_id': 'MISSPELLING',
                    'replacements': [sentence.replace(mistake, correction)]
                })
    return issues

def basic_readability_analysis(text):
    words = text.split()
    sentences = [s for s in text.split('.') if s.strip()]
    if not words or not sentences:
        return {}
    avg_words_per_sentence = len(words) / len(sentences)
    avg_chars_per_word = len(text.replace(' ', '')) / len(words)
    if avg_words_per_sentence <= 10 and avg_chars_per_word <= 4:
        reading_level = "Easy"
    elif avg_words_per_sentence <= 15 and avg_chars_per_word <= 5:
        reading_level = "Medium"
    else:
        reading_level = "Difficult"
    return {
        'reading_level': reading_level,
        'avg_words_per_sentence': round(avg_words_per_sentence, 1),
        'avg_chars_per_word': round(avg_chars_per_word, 1),
        'estimated_grade_level': max(1, min(12, int(avg_words_per_sentence / 2 + avg_chars_per_word)))
    }

def estimate_syllables(text):
    words = text.split()
    total_syllables = 0
    for word in words:
        word = word.lower()
        vowels = 'aeiouy'
        syllable_count = 0
        prev_was_vowel = False
        for char in word:
            if char in vowels:
                if not prev_was_vowel:
                    syllable_count += 1
                prev_was_vowel = True
            else:
                prev_was_vowel = False
        if word.endswith('e') and syllable_count > 1:
            syllable_count -= 1
        total_syllables += max(1, syllable_count)
    return total_syllables

# ... (the remainder of the code consists of all routes, single uploads, report generation, batch processing, and the Flask run block, implemented and indented as per the above conventions from your original script.) ...

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
